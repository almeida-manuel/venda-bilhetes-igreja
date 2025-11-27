import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import sqlite3
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import shutil
import os
import sys
import tkinter.font as tkfont
import json

# Valores de configuração
TICKET_PRICE = 2.0  # preço por bilhete em euros
INITIAL_CASH = 100.0  # caixa inicial em euros

# Config file for persisting settings (preço atual)
def _get_config_dir():
    """Return a writable directory for app config.

    On Windows prefer %APPDATA%\venda-bilhetes-igreja. On other systems use ~/.config/venda-bilhetes-igreja
    """
    try:
        if sys.platform.startswith('win'):
            base = os.getenv('APPDATA') or os.path.expanduser('~')
            cfg_dir = os.path.join(base, 'venda-bilhetes-igreja')
        else:
            base = os.getenv('XDG_CONFIG_HOME') or os.path.join(os.path.expanduser('~'), '.config')
            cfg_dir = os.path.join(base, 'venda-bilhetes-igreja')
        os.makedirs(cfg_dir, exist_ok=True)
        return cfg_dir
    except Exception:
        # fallback to current working directory
        return os.path.abspath('.')

CONFIG_FILE = os.path.join(_get_config_dir(), 'config.json')

def load_config():
    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}

def save_config(cfg: dict):
    try:
        # ensure dir exists
        try:
            os.makedirs(os.path.dirname(CONFIG_FILE), exist_ok=True)
        except Exception:
            pass
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(cfg, f, indent=2, ensure_ascii=False)
        return True
    except Exception:
        return False

# --- Impressão térmica (Windows, impressora USB comum) ---
try:
    import win32print
    import win32ui
    from PIL import Image, ImageWin
    WIN32_AVAILABLE = True
except Exception:
    WIN32_AVAILABLE = False

# Imports opcionais para geração de PDF (reportlab). Se não estiverem presentes,
# funções PDF deverão falhar com uma mensagem amigável.
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus.tables import TableStyle
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

    # Configurações financeiras
    PRICE_PER_TICKET = 2.0  # Euros por bilhete
    STARTING_CASH = 100.0   # Valor inicial em caixa (euros)


def _get_default_printer_name():
    """Retorna o nome da impressora por omissão (Windows) ou None."""
    if not WIN32_AVAILABLE:
        return None
    try:
        return win32print.GetDefaultPrinter()
    except Exception:
        return None


def _send_raw_to_printer(printer_name, data_bytes):
    """Envia bytes raw para a impressora através da API win32print.

    Lança exceções em caso de falha.
    """
    if not WIN32_AVAILABLE:
        raise RuntimeError("pywin32 não está disponível no ambiente.")
    if not printer_name:
        raise ValueError("Nome da impressora não fornecido.")
    # Abrir impressora e enviar trabalho RAW
    hPrinter = None
    try:
        hPrinter = win32print.OpenPrinter(printer_name)
        # StartDocPrinter espera uma tupla (pDocument, pOutputFile, pDatatype)
        # usar RAW para enviar comandos ESC/POS
        win32print.StartDocPrinter(hPrinter, 1, ("Bilhete", None, "RAW"))
        win32print.StartPagePrinter(hPrinter)
        win32print.WritePrinter(hPrinter, data_bytes)
        win32print.EndPagePrinter(hPrinter)
        win32print.EndDocPrinter(hPrinter)
    finally:
        try:
            if hPrinter:
                win32print.ClosePrinter(hPrinter)
        except Exception:
            pass


def imprimir_bilhetes_multiplo_pdf(bilhetes, data_hora, assistente, metodo_pagamento=None, recebido=None, troco=None, quantidade=None, preco=None):
    """Gera um único PDF com uma página por bilhete (80mm largura x altura dinâmica por página).
    Cada bilhete contém: título, imagem.png (se existir) logo a seguir ao título, nº do bilhete,
    data/hora e logo.png (se existir). Depois tenta enviar o PDF para a impressora predefinida.
    """
    if not REPORTLAB_AVAILABLE:
        try:
            messagebox.showwarning(
                "Dependência em Falta",
                "A biblioteca 'reportlab' não está instalada. Instale com: pip install reportlab"
            )
        except Exception:
            print("Dependência em Falta: reportlab")
        return

    try:
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Image, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.utils import ImageReader
    except Exception:
        try:
            messagebox.showwarning("Dependência em Falta", "Falha ao carregar módulos do reportlab.")
        except Exception:
            print("Falha ao carregar módulos do reportlab.")
        return

    import tempfile, time, webbrowser

    # Configurações de página compatíveis com impressora térmica 80mm
    WIDTH_MM = 80
    PAGE_HEIGHT_MM = 127  # altura por página (usar valor dentro do intervalo permitido)
    MARGIN_MM = 1

    width_pt = WIDTH_MM * mm
    height_pt = PAGE_HEIGHT_MM * mm
    margin_pt = MARGIN_MM * mm
    content_width_pt = width_pt - 2 * margin_pt

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleStyle', parent=styles['Normal'],
                                 fontName='Helvetica-Bold', fontSize=12, alignment=TA_CENTER, leading=13)
    small_style = ParagraphStyle('Small', parent=styles['Normal'],
                                 fontName='Helvetica', fontSize=10, alignment=TA_CENTER, leading=11)
    receipt_style = ParagraphStyle('Small', parent=styles['Normal'],
                                 fontName='Helvetica', fontSize=9, alignment=TA_LEFT, leading=11)

    titulo_text = "Bilhete"
    titulo_text2 = "Igreja Nossa Senhora da Oliveira"

    # preparar imagem.png (após o título) e logo.png (no final), calcular tamanhos
    base_dir = os.path.dirname(__file__)
    imagem_path = os.path.join(base_dir, "imagem.png")
    imagem_exists = False
    imagem_w_pt = imagem_h_pt = 0

    logo_path = os.path.join(base_dir, "logo.png")
    logo_exists = False
    logo_w_pt = logo_h_pt = 0

    if os.path.exists(imagem_path):
        try:
            ir = ImageReader(imagem_path)
            iw, ih = ir.getSize()
            if iw > 0:
                # reduzir imagem.png para metade da largura de conteúdo por defeito
                imagem_w_pt = content_width_pt * 0.50
                imagem_h_pt = imagem_w_pt * (ih / float(iw))
                imagem_exists = True
        except Exception:
            imagem_exists = False

    if os.path.exists(logo_path):
        try:
            ir = ImageReader(logo_path)
            iw, ih = ir.getSize()
            if iw > 0:
                # logo ocupa toda a largura de conteúdo por defeito
                logo_w_pt = content_width_pt
                logo_h_pt = logo_w_pt * (ih / float(iw))
                logo_exists = True
        except Exception:
            logo_exists = False

    # criar ficheiro PDF temporário
    tmpdir = tempfile.gettempdir()
    filename = os.path.join(tmpdir, f"bilhetes_batch_{int(time.time())}.pdf")

    story = []
    # Se for uma venda agrupada (quantidade > 1) geramos uma única página
    try:
        if quantidade and int(quantidade) > 1:
            numero = bilhetes[0] if bilhetes else ''
            # logo
            if logo_exists and logo_h_pt > 0:
                try:
                    logo_img = Image(logo_path, width=logo_w_pt, height=logo_h_pt)
                    logo_img.hAlign = 'CENTER'
                    story.append(logo_img)
                except Exception:
                    pass

            # título
            story.append(Paragraph(titulo_text, title_style))
            story.append(Paragraph(titulo_text2, title_style))
            story.append(Spacer(1, 2 * mm))

            # preço total (usar preço fornecido, se houver)
            try:
                unit = float(preco) if preco is not None else TICKET_PRICE
                total_price = float(quantidade) * unit
                story.append(Paragraph(f"Preço total: €{total_price:.2f}", title_style))
            except Exception:
                unit = float(preco) if preco is not None else TICKET_PRICE
                story.append(Paragraph(f"Preço: {unit:.2f}€", title_style))
            story.append(Spacer(1, 2 * mm))

            # imagem.png logo
            if imagem_exists and imagem_h_pt > 0:
                try:
                    img = Image(imagem_path, width=imagem_w_pt, height=imagem_h_pt)
                    img.hAlign = 'CENTER'
                    story.append(img)
                    story.append(Spacer(1, 2 * mm))
                except Exception:
                    pass

            # data/hora e mensagem
            story.append(Paragraph(f"Data/Hora: {data_hora}", small_style))
            story.append(Spacer(1, 2 * mm))

            try:
                story.append(Paragraph(f"Donativo sem contrapartida nos termos do artigo 61 do EBF", small_style))
                story.append(Spacer(1, 2 * mm))
                story.append(Spacer(1, 2 * mm))
            except Exception:
                pass

            # quantidade
            try:
                story.append(Paragraph(f"Quantidade: {int(quantidade)}", receipt_style))
            except Exception:
                pass

            # método de pagamento
            try:
                if metodo_pagamento:
                    story.append(Paragraph(f"Pagamento: {metodo_pagamento}", receipt_style))
                if metodo_pagamento and str(metodo_pagamento).strip().lower() == 'dinheiro':
                    if recebido is not None:
                        story.append(Paragraph(f"Recebido: €{float(recebido):.2f}", receipt_style))
                    if troco is not None:
                        story.append(Paragraph(f"Troco: €{float(troco):.2f}", receipt_style))
                    story.append(Spacer(1, 2 * mm))
            except Exception:
                pass
        else:
            for idx, numero in enumerate(bilhetes):
                # logo
                if logo_exists and logo_h_pt > 0:
                    try:
                        logo_img = Image(logo_path, width=logo_w_pt, height=logo_h_pt)
                        logo_img.hAlign = 'CENTER'
                        story.append(logo_img)
                    except Exception:
                        pass

                # título
                story.append(Paragraph(titulo_text, title_style))
                story.append(Paragraph(titulo_text2, title_style))
                story.append(Spacer(1, 2 * mm))

                # preço (usar preço fornecido, se houver)
                try:
                    unit = float(preco) if preco is not None else TICKET_PRICE
                except Exception:
                    unit = TICKET_PRICE
                story.append(Paragraph(f"Preço: {unit:.2f}€", title_style))
                story.append(Spacer(1, 2 * mm))

                # imagem.png logo
                if imagem_exists and imagem_h_pt > 0:
                    try:
                        img = Image(imagem_path, width=imagem_w_pt, height=imagem_h_pt)
                        img.hAlign = 'CENTER'
                        story.append(img)
                        story.append(Spacer(1, 2 * mm))
                    except Exception:
                        pass

                # data/hora e mensagem
                story.append(Paragraph(f"Data/Hora: {data_hora}", small_style))
                story.append(Spacer(1, 2 * mm))

                try:
                    story.append(Paragraph(f"Donativo sem contrapartida nos termos do artigo 61 do EBF", small_style))
                    story.append(Spacer(1, 2 * mm))
                    story.append(Spacer(1, 2 * mm))
                except Exception:
                    pass

                # método de pagamento
                try:
                    if metodo_pagamento:
                        story.append(Paragraph(f"Pagamento: {metodo_pagamento}", receipt_style))
                    # se pagamento em numerário e valores fornecidos, mostrar recebido e troco
                    if metodo_pagamento and str(metodo_pagamento).strip().lower() == 'dinheiro':
                        if recebido is not None:
                            story.append(Paragraph(f"Recebido: €{float(recebido):.2f}", receipt_style))
                        if troco is not None:
                            story.append(Paragraph(f"Troco: €{float(troco):.2f}", receipt_style))
                        story.append(Spacer(1, 2 * mm))
                except Exception:
                    pass

                # adicionar PageBreak entre bilhetes (não após o último)
                if idx != len(bilhetes) - 1:
                    story.append(PageBreak())
    except Exception:
        # Em caso de erro na formatação do PDF, garantir que continuamos e tentamos gerar o que for possível
        pass

    # gerar PDF com páginas do mesmo tamanho
    doc = SimpleDocTemplate(filename, pagesize=(width_pt, height_pt),
                            leftMargin=margin_pt, rightMargin=margin_pt,
                            topMargin=margin_pt, bottomMargin=margin_pt)
    try:
        doc.build(story)
    except Exception as e:
        try:
            messagebox.showerror("Erro PDF", f"Falha ao criar PDF dos bilhetes:\n{e}")
        except Exception:
            print("Falha ao criar PDF dos bilhetes:", e)
        return

    # tentar imprimir o PDF na impressora predefinida do sistema (Windows preferencialmente)
    try:
        if sys.platform.startswith("win") and WIN32_AVAILABLE:
            try:
                import win32api
                win32api.ShellExecute(0, "print", filename, None, ".", 0)
            except Exception:
                try:
                    os.startfile(filename, "print")
                except Exception:
                    webbrowser.open(filename)
        else:
            try:
                # em sistemas não-Windows, tenta abrir o PDF (usuário imprime manualmente)
                webbrowser.open(filename)
            except Exception:
                pass
    except Exception as e:
        try:
            messagebox.showinfo("PDF Gerado", f"PDF gerado em:\n{filename}\nImpressão automática falhou: {e}")
        except Exception:
            print("PDF gerado em:", filename, "Impressão automática falhou:", e)
# ==========================
# UTILITÁRIOS
# ==========================
def hoje_str():
    return datetime.now().strftime("%Y-%m-%d")

# Accessibility: increase font sizes for better readability
FONT_INCREASE = 2  # change this number to increase/decrease size
def AF(size, *opts):
    # returns a font tuple with increased size and any extra options (e.g., 'bold')
    try:
        sz = int(size) + FONT_INCREASE
    except Exception:
        sz = size
    return ("Segoe UI", sz) + tuple(opts)


def agora_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# Increase common Tk named fonts for accessibility. Call after creating a Tk root.
def adjust_fonts(delta=FONT_INCREASE):
    names = [
        "TkDefaultFont", "TkTextFont", "TkMenuFont", "TkHeadingFont",
        "TkCaptionFont", "TkSmallCaptionFont", "TkIconFont", "TkTooltipFont"
    ]
    for n in names:
        try:
            f = tkfont.nametofont(n)
            # some fonts report negative sizes on Windows; handle gracefully
            try:
                current = int(f['size'])
            except Exception:
                current = f['size']
            try:
                f.configure(size=current + delta)
            except Exception:
                pass
        except Exception:
            pass


# ==========================
# GESTOR DE BASE DE DADOS
# ==========================
class DatabaseManager:
    def __init__(self, path="bilhetes.db"):
        self.path = path
        self.conn = sqlite3.connect(self.path, detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES)
        self.cursor = self.conn.cursor()
        self._criar_tabela()

    def _criar_tabela(self):
        # Cria tabela com coluna 'anotacoes' (opcional). Se a tabela já existir sem a coluna,
        # fazemos uma migração simples adicionando a coluna.
        # incluir coluna 'preco' para armazenar o preço unitário do bilhete ao registar
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS registos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                data_hora TEXT,
                assistente TEXT,
                nacionalidade TEXT,
                numero_bilhete TEXT,
                metodo_pagamento TEXT,
                fatura TEXT,
                contribuinte TEXT,
                preco REAL,
                anotacoes TEXT
            )
        """)
        self.conn.commit()

        # Verificar se a coluna 'anotacoes' existe; se não, adicioná-la (migração para versões antigas)
        try:
            self.cursor.execute("PRAGMA table_info(registos)")
            cols = [r[1] for r in self.cursor.fetchall()]
            # adicionar colunas ausentes por migração simples
            if 'anotacoes' not in cols:
                self.cursor.execute("ALTER TABLE registos ADD COLUMN anotacoes TEXT")
                self.conn.commit()
                cols.append('anotacoes')
            if 'preco' not in cols:
                try:
                    self.cursor.execute("ALTER TABLE registos ADD COLUMN preco REAL")
                    self.conn.commit()
                except Exception:
                    pass
            # Se houver linhas sem preco (migração de versões antigas), preencher com preço padrão do config ou constante
            try:
                cfg = load_config()
                default_price = float(cfg.get('ticket_price', TICKET_PRICE))
            except Exception:
                default_price = TICKET_PRICE
            try:
                self.cursor.execute("UPDATE registos SET preco = ? WHERE preco IS NULL", (default_price,))
                self.conn.commit()
            except Exception:
                pass
        except Exception:
            # Se qualquer erro ocorrer aqui, não queremos quebrar a inicialização; seguir em frente
            pass

        # tabela para eventos (registos auxiliares como 'nao_entraram')
        try:
            self.cursor.execute("""
                CREATE TABLE IF NOT EXISTS eventos (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    timestamp TEXT,
                    event_type TEXT,
                    count INTEGER,
                    assistente TEXT,
                    notes TEXT
                )
            """)
            self.conn.commit()
        except Exception:
            pass

    def inserir_evento(self, event_type, count=None, assistente=None, notes=None, timestamp=None):
        try:
            ts = timestamp if timestamp is not None else datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.cursor.execute(
                "INSERT INTO eventos (timestamp, event_type, count, assistente, notes) VALUES (?, ?, ?, ?, ?)",
                (ts, event_type, count, assistente, notes)
            )
            self.conn.commit()
        except Exception:
            pass

    def obter_eventos_por_tipo(self, event_type, dia_str=None):
        if dia_str is None:
            dia_str = hoje_str()
        try:
            self.cursor.execute(
                "SELECT id, timestamp, event_type, count, assistente, notes FROM eventos WHERE event_type = ? AND date(timestamp) = ? ORDER BY id DESC",
                (event_type, dia_str)
            )
            return self.cursor.fetchall()
        except Exception:
            return []

    def apagar_evento_por_id(self, event_id):
        try:
            self.cursor.execute("DELETE FROM eventos WHERE id = ?", (event_id,))
            self.conn.commit()
            return True
        except Exception:
            return False

    def inserir_registo(self, data_hora, assistente, nacionalidade, numero_bilhete, metodo_pagamento, fatura, contribuinte, anotacoes=None):
        # manter compatibilidade: aceitar um parametro opcional 'preco' via kwargs se fornecido
        preco = None
        # detectar se 'anotacoes' foi passado como positional (legacy) ou se foi fornecido preço via keyword
        # chamada típica: inserir_registo(..., anotacoes)
        # Para chamadas internas novas, passamos preco como a última positional ou via keyword
        try:
            # tentar obter de self if foi anexado temporariamente (não ideal, mas compatível)
            preco = getattr(self, '_pending_preco', None)
        except Exception:
            preco = None
        # executar insert incluindo preco
        self.cursor.execute("""
            INSERT INTO registos (data_hora, assistente, nacionalidade, numero_bilhete, metodo_pagamento, fatura, contribuinte, preco, anotacoes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (data_hora, assistente, nacionalidade, numero_bilhete, metodo_pagamento, fatura, contribuinte, preco, anotacoes))
        self.conn.commit()

    def atualizar_anotacoes_por_numero(self, numero_bilhete, novo_texto):
        """Anexa (ou define) o texto de anotacoes para o registo mais recente com o numero_bilhete.

        Retorna True se actualizado com sucesso, False caso contrario.
        """
        try:
            self.cursor.execute("SELECT id, anotacoes FROM registos WHERE numero_bilhete = ? ORDER BY id DESC LIMIT 1", (numero_bilhete,))
            row = self.cursor.fetchone()
            if not row:
                return False
            rid, existing = row
            if existing and existing.strip():
                combinado = f"{existing} | {novo_texto}"
            else:
                combinado = novo_texto
            self.cursor.execute("UPDATE registos SET anotacoes = ? WHERE id = ?", (combinado, rid))
            self.conn.commit()
            return True
        except Exception:
            return False

    def ultimo_numero_bilhete(self):
        self.cursor.execute("SELECT numero_bilhete FROM registos ORDER BY id DESC LIMIT 1")
        row = self.cursor.fetchone()
        return row[0] if row else None

    def obter_registos_do_dia(self, dia_str=None):
        if dia_str is None:
            dia_str = hoje_str()
        # Assumimos data_hora armazenada como 'YYYY-MM-DD HH:MM:SS'
        self.cursor.execute("""
            SELECT data_hora, assistente, nacionalidade, numero_bilhete, metodo_pagamento, fatura, contribuinte, preco, anotacoes
            FROM registos
            WHERE date(data_hora) = ?
            ORDER BY id DESC
        """, (dia_str,))
        return self.cursor.fetchall()

    def procurar_por_bilhete(self, termo):
        termo_like = f"%{termo}%"
        self.cursor.execute("""
            SELECT data_hora, assistente, nacionalidade, numero_bilhete, metodo_pagamento, fatura, contribuinte, preco, anotacoes
            FROM registos
            WHERE numero_bilhete LIKE ?
            ORDER BY id DESC
        """, (termo_like,))
        return self.cursor.fetchall()

    def fechar(self):
        try:
            self.conn.close()
        except Exception:
            pass


# ==========================
# INTERFACE - LOGIN
# ==========================
class JanelaLogin:
    def __init__(self):
        self.login = tk.Tk()
        # aplicar ajuste de fontes para acessibilidade
        adjust_fonts()
        self.login.title("Login de Assistente")
        self.login.geometry("400x320")
        self.login.resizable(False, False)
        self.login.configure(bg="#f8fafc")
        
        # Centralizar na tela
        self.login.eval('tk::PlaceWindow . center')

        # Container principal
        main_frame = tk.Frame(self.login, bg="#f8fafc", padx=20, pady=30)
        main_frame.pack(expand=True, fill="both")

        # Título
        title_frame = tk.Frame(main_frame, bg="#f8fafc")
        title_frame.pack(pady=(0, 25))
        
        tk.Label(title_frame, text="Venda de Bilhetes", font=("Segoe UI", 20, "bold"), 
                bg="#f8fafc", fg="#2d3748").pack()
        tk.Label(title_frame, text="Sistema de Gestão", font=("Segoe UI", 12), 
                bg="#f8fafc", fg="#718096").pack(pady=(5, 0))

        # Form container
        form_frame = tk.Frame(main_frame, bg="#f8fafc")
        form_frame.pack(fill="x", pady=20)

        tk.Label(form_frame, text="Nome do Assistente:", font=("Segoe UI", 11, "bold"), 
                bg="#f8fafc", fg="#4a5568").pack(anchor="w", pady=(0, 8))

        self.entry_nome = ttk.Entry(form_frame, font=("Segoe UI", 11), width=25)
        self.entry_nome.pack(fill="x", pady=(0, 20))
        self.entry_nome.focus()

        # Bind Enter key to login
        self.entry_nome.bind('<Return>', lambda e: self.confirmar())

        # Botão de entrar
        btn_entrar = tk.Button(form_frame, text="Entrar no Sistema", 
                              font=("Segoe UI", 11, "bold"),
                              bg="#4299e1", fg="white",
                              activebackground="#3182ce",
                              activeforeground="white",
                              relief="flat",
                              padx=10, pady=10,
                              command=self.confirmar)
        btn_entrar.pack(pady=0)
        
        # Efeito hover no botão
        def on_enter(e):
            btn_entrar['background'] = '#3182ce'
        def on_leave(e):
            btn_entrar['background'] = '#4299e1'
        btn_entrar.bind("<Enter>", on_enter)
        btn_entrar.bind("<Leave>", on_leave)

        # Footer
        footer_frame = tk.Frame(main_frame, bg="#f8fafc")
        footer_frame.pack(side="bottom", pady=10)
        tk.Label(footer_frame, text="Digite seu nome para acessar o sistema", 
                font=("Segoe UI", 9), bg="#f8fafc", fg="#a0aec0").pack()

        self.login.mainloop()

    def confirmar(self):
        nome = self.entry_nome.get().strip()
        if not nome:
            messagebox.showwarning("Aviso", "Insira o nome do assistente.")
            return
        self.login.destroy()
        JanelaPrincipal(nome)


# ==========================
# INTERFACE - APLICAÇÃO PRINCIPAL
# ==========================
class JanelaPrincipal:
    def __init__(self, assistente_nome):
        self.assistente = assistente_nome
        self.dia_fechado = False

        # carregar configuração (preço persistido)
        try:
            cfg = load_config()
            self.ticket_price = float(cfg.get('ticket_price', TICKET_PRICE))
        except Exception:
            self.ticket_price = TICKET_PRICE

        # DB
        try:
            self.db = DatabaseManager("bilhetes.db")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao abrir BD: {e}")
            return

        # Janela principal
        self.root = tk.Tk()
        # aplicar ajuste de fontes para acessibilidade
        adjust_fonts()
        self.root.title(f"Venda de Bilhetes - Assistente: {self.assistente}")
        self.root.geometry("1200x750")
        self.root.resizable(True, True)
        self.root.configure(bg="#f8fafc")
        
        # Centralizar na tela
        self.root.eval('tk::PlaceWindow . center')

        # Criar UI
        self._criar_interface()
        self.atualizar_tabela()
        self._atualizar_status()

        # Fechar corretamente
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        self.root.mainloop()

    # --------------------------
    # INTERFACE
    # --------------------------
    def _criar_interface(self):
        # Header
        header = tk.Frame(self.root, bg="#2d3748", height=80)
        header.pack(fill="x", side="top")
        header.pack_propagate(False)

        # Conteúdo do header
        header_content = tk.Frame(header, bg="#2d3748")
        header_content.pack(expand=True, fill="both", padx=30)

        tk.Label(header_content, text="Sistema de Venda de Bilhetes",
                 font=AF(16, "bold"), bg="#2d3748", fg="white").pack(side="left")

        user_frame = tk.Frame(header_content, bg="#2d3748")
        user_frame.pack(side="right")

        tk.Label(user_frame, text="Assistente:", font=AF(10),
                 bg="#2d3748", fg="#cbd5e0").pack(side="left")
        self.lbl_assistente = tk.Label(user_frame, text=self.assistente,
                                      font=AF(10, "bold"),
                                      bg="#2d3748", fg="white")
        self.lbl_assistente.pack(side="left", padx=(5, 15))

        # Botão trocar assistente com estilo moderno
        btn_trocar = tk.Button(user_frame, text="Trocar Assistente",
                               font=AF(9),
                               bg="#4a5568", fg="white",
                               activebackground="#718096",
                               activeforeground="white",
                               relief="flat",
                               padx=12, pady=4,
                               command=self._popup_trocar_assistente)
        btn_trocar.pack(side="left")

        # Botão único para registar entrada/saída do organista (toggle)
        self.btn_organista = tk.Button(user_frame, text="Registar Entrada Organista",
                                       font=AF(9),
                                       bg="#48bb78", fg="white",
                                       activebackground="#38a169",
                                       activeforeground="white",
                                       relief="flat",
                                       padx=10, pady=4,
                                       command=self._on_click_organista_toggle)
        self.btn_organista.pack(side="left", padx=(8, 4))

        # Container principal
        # Container principal com scroll (garante que todo o conteúdo fica acessível em ecrãs pequenos)
        container_outer = tk.Frame(self.root)
        container_outer.pack(expand=True, fill="both", padx=20, pady=20)

        canvas = tk.Canvas(container_outer, bg="#f8fafc", highlightthickness=0)
        vscroll = ttk.Scrollbar(container_outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vscroll.set)
        vscroll.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # frame onde será colocado todo o conteúdo (usamos o mesmo nome main_container para compatibilidade)
        main_container = tk.Frame(canvas, bg="#f8fafc")
        self._canvas_window = canvas.create_window((0, 0), window=main_container, anchor="nw")

        # atualizar scrollregion quando o conteúdo mudar e forçar largura interna igual à do canvas
        def _on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        def _on_canvas_configure(event):
            canvas.itemconfig(self._canvas_window, width=event.width)
        main_container.bind("<Configure>", _on_frame_configure)
        canvas.bind("<Configure>", _on_canvas_configure)

        # suporte a roda do rato (Windows/Mac/Linux)
        def _on_mousewheel(event):
            try:
                if sys.platform == "darwin":
                    delta = -1 * int(event.delta)
                else:
                    delta = -1 * int(event.delta / 120)
            except Exception:
                delta = 0
            if delta:
                canvas.yview_scroll(delta, "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        canvas.bind_all("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))
        canvas.bind_all("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))

        # Painel esquerdo - Formulário de venda
        left_panel = tk.Frame(main_container, bg="white", relief="flat", bd=1)
        left_panel.pack(side="left", fill="y", padx=(0, 15))

        # Título do painel
        panel_title = tk.Frame(left_panel, bg="#4299e1", height=40)
        panel_title.pack(fill="x", side="top")
        panel_title.pack_propagate(False)
        tk.Label(panel_title, text="Nova Venda", font=AF(12, "bold"), bg="#4299e1", fg="white").pack(expand=True)

        # Form container
        form_container = tk.Frame(left_panel, bg="white", padx=20, pady=20)
        form_container.pack(expand=True, fill="both")

        # Campos do formulário
        fields = [
            ("Nacionalidade:", "combo_nacionalidade"),
            ("Método de Pagamento:", "combo_pagamento"),
            ("Recibo:", "combo_fatura"),
            ("Anotações (opcional):", "entry_anotacoes"),
            ("Quantidade:", "spin_quantidade")
        ]

        nacionalidades = [
            "Português", "Brasileiro", "Espanhol", "Inglês", "Francês", "Italiano", "Asiático", "Alemão", "Outros"
        ]

        # usamos um contador de linhas para poder inserir a entrada manual logo abaixo do combo de nacionalidade
        line = 0
        for label, field_type in fields:
            tk.Label(form_container, text=label, font=AF(10, "bold"), bg="white", fg="#4a5568").grid(row=line, column=0, sticky="w", pady=12, padx=(0, 10))

            if "combo" in field_type:
                values = nacionalidades if "nacionalidade" in field_type else ["Dinheiro", "Cartão"] if "pagamento" in field_type else ["Sim", "Não"]
                default = "Português" if "nacionalidade" in field_type else "Dinheiro" if "pagamento" in field_type else "Não"
                combo = ttk.Combobox(form_container, values=values, state="readonly", font=AF(10), width=22)
                combo.set(default)
                combo.grid(row=line, column=1, sticky="ew", pady=12)
                setattr(self, field_type, combo)
                # Se este for o combo de nacionalidade, ligar o evento para mostrar entrada manual
                if "nacionalidade" in field_type:
                    # a entrada manual deve ficar imediatamente abaixo deste combo
                    self._manual_row = line + 1
                    combo.bind('<<ComboboxSelected>>', lambda e: self._on_nacionalidade_change(e))
                    # reservar linha para a entrada manual (não inserida ainda)
                    line += 1
                # Se este for o combo de fatura, reservar linha para o Nº Contribuinte
                if "fatura" in field_type:
                    self._contrib_row = line + 1
                    combo.bind('<<ComboboxSelected>>', lambda e: self._on_fatura_change(e))
                    # reservar linha para a entrada de contribuinte
                    line += 1
            elif "entry" in field_type:
                # Campo de anotações: usar um widget multi-linha (ScrolledText)
                if "anotacoes" in field_type:
                    txt = scrolledtext.ScrolledText(form_container, font=AF(10), width=36, height=4, wrap=tk.WORD)
                    txt.grid(row=line, column=1, sticky="ew", pady=12)
                    setattr(self, field_type, txt)
                else:
                    entry = ttk.Entry(form_container, font=AF(10), width=25)
                    entry.grid(row=line, column=1, sticky="ew", pady=12)
                    setattr(self, field_type, entry)
            elif "spin" in field_type:
                spin = ttk.Spinbox(form_container, from_=1, to=50, width=10, font=AF(10))
                # por padrão, 0 pessoas não entraram
                if 'nao_entraram' in field_type:
                    spin.set(0)
                else:
                    spin.set(1)
                spin.grid(row=line, column=1, sticky="w", pady=12)
                setattr(self, field_type, spin)

            # avançar para a próxima linha
            line += 1

        # criar placeholder para entrada manual (inicialmente escondido)
        self.manual_nacionalidade_var = tk.StringVar()
        self.entry_manual_nacionalidade = ttk.Entry(form_container, textvariable=self.manual_nacionalidade_var, font=AF(10), width=25)
        # criar também rótulo explícito (não gridado ainda)
        self.lbl_manual_nacionalidade = tk.Label(form_container, text="Especificar Nacionalidade:", font=AF(10, "bold"), bg="white", fg="#4a5568")
        # não grid ainda; será exibida quando necessário via _on_nacionalidade_change

        # placeholder para contribuinte (aparece apenas se Fatura == 'Sim')
        self.contribuinte_var = tk.StringVar()
        self.entry_contribuinte = ttk.Entry(form_container, textvariable=self.contribuinte_var, font=AF(10), width=25)
        self.lbl_contribuinte = tk.Label(form_container, text="Nº Contribuinte:", font=AF(10, "bold"), bg="white", fg="#4a5568")


        # Botões de ação
        btn_frame = tk.Frame(form_container, bg="white")
        # colocar os botões abaixo de todos os campos (usar 'line' para ficar abaixo da Quantidade)
        btn_frame.grid(row=line, column=0, columnspan=2, pady=(25, 10))

        # Botão Guardar
        btn_guardar = tk.Button(btn_frame, text="✓ Guardar Registo",
                               font=AF(11, "bold"),
                               bg="#48bb78", fg="white",
                               activebackground="#38a169",
                               activeforeground="white",
                               relief="flat",
                               padx=20, pady=10,
                               command=self.guardar_registo)
        btn_guardar.pack(side="left", padx=(0, 10))

        # Botão Fechar Dia
        btn_fechar = tk.Button(btn_frame, text="📊 Fechar o Dia",
                              font=AF(11, "bold"),
                              bg="#ed8936", fg="white",
                              activebackground="#dd6b20",
                              activeforeground="white",
                              relief="flat",
                              padx=20, pady=10,
                              command=self.fechar_dia)
        btn_fechar.pack(side="left")

        # Linha abaixo dos botões: registo de pessoas que não entraram
        try:
            reg_row = line + 1
            tk.Label(form_container, text="Registar Não Entram:", font=AF(10, "bold"), bg="white", fg="#4a5568").grid(row=reg_row, column=0, sticky="w", pady=8, padx=(0,10))
            self.spin_reg_nao_entraram = ttk.Spinbox(form_container, from_=1, to=50, width=10, font=AF(10))
            self.spin_reg_nao_entraram.set(1)
            self.spin_reg_nao_entraram.grid(row=reg_row, column=1, sticky="w", pady=8)
            # botão para registar o(s) não entrado(s)
            def _on_registar_nao_entraram():
                # impedir registos após o dia estar fechado
                try:
                    if getattr(self, 'dia_fechado', False):
                        messagebox.showwarning("Aviso", "O dia já está fechado. Não é possível registar mais 'Não Entram'.")
                        return
                except Exception:
                    pass
                try:
                    cnt = int(self.spin_reg_nao_entraram.get())
                except Exception:
                    messagebox.showwarning("Aviso", "Quantidade inválida para registo de Não Entram.")
                    return
                if cnt <= 0:
                    messagebox.showwarning("Aviso", "Quantidade deve ser pelo menos 1.")
                    return
                # Registar um único evento com a quantidade (count) e o timestamp atual.
                ts = agora_str()
                try:
                    self.db.inserir_evento('nao_entraram', count=cnt, assistente=self.assistente, notes=None, timestamp=ts)
                    # mostrar apenas a hora ao utilizador
                    try:
                        hora = ts.split(' ')[1]
                    except Exception:
                        hora = ts
                    messagebox.showinfo("Registo Efetuado", f"Registados {cnt} não entrado(s) às {hora}")
                    self._set_status(f"{cnt} 'Não Entraram' registado(s) às {hora}.")
                    try:
                        # resetar spinbox para 1 após registo
                        self.spin_reg_nao_entraram.set(1)
                    except Exception:
                        pass
                except Exception:
                    messagebox.showwarning("Aviso", "Falha ao registar Não Entram.")

            btn_reg_nao = tk.Button(form_container, text="❌ Registar Não Entram", font=AF(10), bg="#f56565", fg="white", activebackground="#c53030", relief="flat", command=_on_registar_nao_entraram)
            # colocar o botão abaixo da spinbox (na próxima linha, mesma coluna da spinbox)
            btn_reg_nao.grid(row=reg_row+1, column=1, sticky="w", pady=(4, 0))
            # expor como atributo para poder ser desativado ao fechar o dia
            try:
                self.btn_reg_nao = btn_reg_nao
            except Exception:
                pass
        except Exception:
            # se algo falhar aqui, não quebrar a criação da interface
            pass

        # Painel direito - Estatísticas e dados
        right_panel = tk.Frame(main_container, bg="#f8fafc")
        right_panel.pack(side="left", expand=True, fill="both")

        # Frame de estatísticas
        stats_frame = tk.Frame(right_panel, bg="white", relief="flat", bd=1)
        stats_frame.pack(fill="x", pady=(0, 15))

        stats_title = tk.Frame(stats_frame, bg="#38a169", height=40)
        stats_title.pack(fill="x", side="top")
        stats_title.pack_propagate(False)
        tk.Label(stats_title, text="Estatísticas do Dia", font=AF(12, "bold"), bg="#38a169", fg="white").pack(expand=True)

        stats_content = tk.Frame(stats_frame, bg="white", padx=20, pady=15)
        stats_content.pack(expand=True, fill="both")

        self.lbl_total_today = tk.Label(stats_content, text="Total de bilhetes hoje: 0", font=AF(12, "bold"), bg="white", fg="#2d3748")
        self.lbl_total_today.pack(anchor="w", pady=(0, 15))

        # Caixa / Totais de pagamento
        money_frame = tk.Frame(stats_content, bg="white")
        money_frame.pack(anchor="w", pady=(0, 10))

        self.lbl_caixa_total = tk.Label(money_frame, text=f"Caixa total: €{INITIAL_CASH:.2f}", font=AF(10, "bold"), bg="white", fg="#2d3748")
        self.lbl_caixa_total.pack(anchor="w")
        self.lbl_numerario = tk.Label(money_frame, text="Numerário: €0.00", font=AF(10), bg="white", fg="#2d3748")
        self.lbl_numerario.pack(anchor="w")
        self.lbl_multibanco = tk.Label(money_frame, text="Multibanco: €0.00", font=AF(10), bg="white", fg="#2d3748")
        self.lbl_multibanco.pack(anchor="w")

        # Tabela de nacionalidades
        nat_frame = tk.Frame(stats_content, bg="white")
        nat_frame.pack(fill="both", expand=True)

        tk.Label(nat_frame, text="Distribuição por Nacionalidade:", font=AF(10, "bold"), bg="white", fg="#4a5568").pack(anchor="w")

        self.lst_nacionalidades = ttk.Treeview(nat_frame, columns=("nacionalidade", "count"), show="headings", height=8)
        self.lst_nacionalidades.heading("nacionalidade", text="Nacionalidade")
        self.lst_nacionalidades.heading("count", text="Total")
        self.lst_nacionalidades.column("nacionalidade", width=180)
        self.lst_nacionalidades.column("count", width=80, anchor="center")
        
        # Scrollbar para a treeview
        nat_scroll = ttk.Scrollbar(nat_frame, orient="vertical", command=self.lst_nacionalidades.yview)
        self.lst_nacionalidades.configure(yscrollcommand=nat_scroll.set)
        
        self.lst_nacionalidades.pack(side="left", fill="both", expand=True, pady=(8, 0))
        nat_scroll.pack(side="right", fill="y")

        # Frame de pesquisa
        search_frame = tk.Frame(right_panel, bg="white", relief="flat", bd=1)
        search_frame.pack(fill="x", pady=(0, 15))

        search_content = tk.Frame(search_frame, bg="white", padx=20, pady=15)
        search_content.pack(expand=True, fill="both")

        tk.Label(search_content, text="Pesquisar Bilhetes", font=AF(11, "bold"), bg="white", fg="#4a5568").pack(anchor="w", pady=(0, 10))

        search_controls = tk.Frame(search_content, bg="white")
        search_controls.pack(fill="x")

        tk.Label(search_controls, text="Nº do bilhete:", font=AF(10), bg="white", fg="#4a5568").pack(side="left")

        self.entry_search = ttk.Entry(search_controls, font=AF(10), width=20)
        self.entry_search.pack(side="left", padx=8)

        btn_pesquisar = tk.Button(search_controls, text="🔍 Pesquisar",
                                 font=AF(9),
                                 bg="#4299e1", fg="white",
                                 activebackground="#3182ce",
                                 activeforeground="white",
                                 relief="flat",
                                 padx=12, pady=4,
                                 command=self.pesquisar_bilhete)
        btn_pesquisar.pack(side="left", padx=(5, 10))

        btn_limpar = tk.Button(search_controls, text="Limpar Filtro",
                              font=AF(9),
                              bg="#a0aec0", fg="white",
                              activebackground="#718096",
                              activeforeground="white",
                              relief="flat",
                              padx=12, pady=4,
                              command=self.atualizar_tabela)
        btn_limpar.pack(side="left")

        # Tabela principal de registos
        table_frame = tk.Frame(right_panel, bg="white", relief="flat", bd=1)
        table_frame.pack(expand=True, fill="both")

        table_title = tk.Frame(table_frame, bg="#4a5568", height=40)
        table_title.pack(fill="x", side="top")
        table_title.pack_propagate(False)
        tk.Label(table_title, text="Registos de Hoje", font=AF(12, "bold"), bg="#4a5568", fg="white").pack(expand=True)

        table_content = tk.Frame(table_frame, bg="white")
        table_content.pack(expand=True, fill="both", padx=2, pady=2)

        # Treeview (registos do dia) - incluir coluna 'preco' (não remove a coluna 'anotacoes' que fica por fim)
        cols = ("data_hora", "assistente", "nacionalidade", "numero_bilhete", "metodo_pagamento", "fatura", "contribuinte", "preco", "anotacoes")
        self.tree = ttk.Treeview(table_content, columns=cols, show="headings", height=15)
        
        # Configurar colunas
        # Map some internal column names to nicer display headings
        display_names = {
            'fatura': 'Recibo',
            'preco': 'Preço (€)'
        }
        for c in cols:
            heading = display_names.get(c, c.replace("_", " ").capitalize())
            self.tree.heading(c, text=heading)
            # aumentar largura da coluna 'anotacoes'
            col_width = 220 if c == 'anotacoes' else 90 if c == 'preco' else 120
            self.tree.column(c, width=col_width, anchor="center")
        
        # Scrollbars
        v_scroll = ttk.Scrollbar(table_content, orient="vertical", command=self.tree.yview)
        h_scroll = ttk.Scrollbar(table_content, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
        
        # Usar grid para garantir que a scrollbar horizontal fique sempre visível
        # e que os componentes redimensionem corretamente dentro do frame
        table_content.grid_rowconfigure(0, weight=1)
        table_content.grid_columnconfigure(0, weight=1)
        self.tree.grid(row=0, column=0, sticky='nsew')
        v_scroll.grid(row=0, column=1, sticky='ns')
        h_scroll.grid(row=1, column=0, columnspan=2, sticky='ew')

        # Bind duplo-clique para mostrar detalhes (anotações completas)
        self.tree.bind('<Double-1>', self._mostrar_detalhes)

        # Permitir scroll horizontal usando Shift + roda do rato (melhora usabilidade em colunas largas)
        def _on_shift_mousewheel(event):
            try:
                if sys.platform == "darwin":
                    delta = -1 * int(event.delta)
                else:
                    delta = -1 * int(event.delta / 120)
            except Exception:
                delta = 0
            if delta:
                try:
                    self.tree.xview_scroll(delta, "units")
                except Exception:
                    pass

        # Bind globalmente para funcionar mesmo quando o foco não está exactamente na tree
        self.tree.bind_all("<Shift-MouseWheel>", _on_shift_mousewheel)
        # Suporte para sistemas que usam Button-4/Button-5
        self.tree.bind_all("<Shift-Button-4>", lambda e: self.tree.xview_scroll(-1, "units"))
        self.tree.bind_all("<Shift-Button-5>", lambda e: self.tree.xview_scroll(1, "units"))

        # Status bar
        status_bar = tk.Frame(self.root, bg="#e2e8f0", height=30)
        status_bar.pack(fill="x", side="bottom")
        status_bar.pack_propagate(False)
        
        self.status_var = tk.StringVar()
        self.status_var.set("Sistema pronto - Aguardando ações")
        status_label = tk.Label(status_bar, textvariable=self.status_var, font=AF(9), bg="#e2e8f0", fg="#4a5568")
        status_label.pack(side="left", padx=15)
        # Atualizar estado do botão do organista (entrada/saída) conforme eventos existentes para hoje
        try:
            self._update_organista_button_state()
        except Exception:
            pass

        # Atalho para alterar o preço dos bilhetes: Ctrl+Shift+P
        try:
            # vincular tanto P maiúsculo como p minúsculo para compatibilidade
            self.root.bind_all('<Control-Shift-P>', lambda e: self._on_shortcut_change_price())
            self.root.bind_all('<Control-Shift-p>', lambda e: self._on_shortcut_change_price())
        except Exception:
            pass

    # --------------------------
    # FUNÇÕES DE AÇÃO
    # --------------------------
    def _popup_trocar_assistente(self):
        if self.dia_fechado:
            messagebox.showwarning("Aviso", "O dia já está fechado! Não é possível trocar assistente.")
            return
        popup = tk.Toplevel(self.root)
        popup.title("Trocar Assistente")
        popup.geometry("320x140")
        popup.resizable(False, False)
        popup.configure(bg="#f8fafc")
        popup.transient(self.root)
        popup.grab_set()
        
        ttk.Label(popup, text="Novo nome do assistente:").pack(pady=(12, 6))
        entry = ttk.Entry(popup, width=30)
        entry.pack(pady=(0, 10))
        entry.focus()

        def confirmar():
            nome = entry.get().strip()
            if not nome:
                messagebox.showwarning("Aviso", "Insira um nome válido.")
                return
            self.assistente = nome
            self.lbl_assistente.config(text=self.assistente)
            self.root.title(f"Venda de Bilhetes - Assistente: {self.assistente}")
            popup.destroy()
            self._set_status(f"Assistente alterado para: {self.assistente}")

        ttk.Button(popup, text="Confirmar", command=confirmar).pack(pady=(0, 10))

    def _popup_registrar_organista(self, event_type):
        """Mostra popup para pedir o nome do organista (opcional nota) e grava evento ('organista_entrada' ou 'organista_saida')."""
        if self.dia_fechado:
            messagebox.showwarning("Aviso", "O dia já está fechado. Não é possível registar eventos.")
            return
        popup = tk.Toplevel(self.root)
        popup.title("Registo Organista")
        popup.geometry("380x180")
        popup.transient(self.root)
        popup.grab_set()

        tk.Label(popup, text="Nome do Organista:", font=AF(10)).pack(anchor='w', padx=12, pady=(10, 2))
        nome_var = tk.StringVar()
        nome_entry = ttk.Entry(popup, textvariable=nome_var, width=40)
        nome_entry.pack(padx=12)
        nome_entry.focus()

        tk.Label(popup, text="Notas (opcional):", font=AF(10)).pack(anchor='w', padx=12, pady=(8, 2))
        notas_var = tk.StringVar()
        notas_entry = ttk.Entry(popup, textvariable=notas_var, width=40)
        notas_entry.pack(padx=12)

        def confirmar():
            organista = nome_var.get().strip() or "Organista"
            notas = notas_var.get().strip() or None
            ts = agora_str()
            # juntar nome e notas num único campo notes (se houver notas, separar por '|')
            if notas:
                notes_field = f"{organista}|{notas}"
            else:
                notes_field = organista
            try:
                self.db.inserir_evento(event_type, count=None, assistente=self.assistente, notes=notes_field, timestamp=ts)
                try:
                    hora = ts.split(' ')[1]
                except Exception:
                    hora = ts
                tipo = 'Entrada' if event_type == 'organista_entrada' else 'Saída'
                messagebox.showinfo("Registo Efetuado", f"{tipo} do organista '{organista}' registada às {hora}")
                self._set_status(f"Organista {tipo.lower()} registada: {organista} às {hora}")
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao registar evento do organista: {e}")
            finally:
                try:
                    popup.destroy()
                except Exception:
                    pass
                try:
                    # atualizar o texto do botão para refletir estado atual (entrada vs saída)
                    self._update_organista_button_state()
                except Exception:
                    pass

        btns = tk.Frame(popup)
        btns.pack(pady=(12, 8))
        ttk.Button(btns, text="Confirmar", command=confirmar).pack(side='left', padx=8)
        ttk.Button(btns, text="Cancelar", command=popup.destroy).pack(side='left')

    def _on_shortcut_change_price(self, event=None):
        """Abre um popup modal para alterar o preço por bilhete via atalho Ctrl+Shift+P.

        Valida o valor (float > 0), atualiza a variável global TICKET_PRICE e recalcula as estatísticas.
        """
        try:
            if getattr(self, 'dia_fechado', False):
                messagebox.showwarning("Aviso", "O dia já está fechado. Não é possível alterar o preço.")
                return
        except Exception:
            pass

        popup = tk.Toplevel(self.root)
        popup.title("Alterar Preço do Bilhete")
        popup.geometry("360x140")
        popup.transient(self.root)
        popup.grab_set()

        tk.Label(popup, text="Novo preço por bilhete (€):", font=AF(10)).pack(anchor='w', padx=12, pady=(12, 6))
        price_var = tk.StringVar()
        # preencher com o preço atual (valor carregado em self.ticket_price)
        try:
            price_var.set(str(float(getattr(self, 'ticket_price', TICKET_PRICE))))
        except Exception:
            price_var.set('0.00')

        entry = ttk.Entry(popup, textvariable=price_var, width=20, font=AF(10))
        entry.pack(padx=12)
        entry.focus()

        def confirmar():
            s = price_var.get().strip().replace(',', '.')
            try:
                val = float(s)
                if val <= 0:
                    raise ValueError()
            except Exception:
                messagebox.showwarning("Aviso", "Introduza um preço válido (ex.: 2.00).")
                return
            # actualizar preço atual em memória e persistir no ficheiro de configuração
            try:
                self.ticket_price = val
                cfg = load_config()
                cfg['ticket_price'] = val
                save_config(cfg)
            except Exception:
                pass
            try:
                # recalcular e atualizar a UI que depende do preço
                self._atualizar_estatisticas()
                self._set_status(f"Preço por bilhete alterado para €{val:.2f}")
            except Exception:
                pass
            popup.destroy()

        btns = tk.Frame(popup)
        btns.pack(pady=(10, 8))
        ttk.Button(btns, text="Confirmar", command=confirmar).pack(side='left', padx=8)
        ttk.Button(btns, text="Cancelar", command=popup.destroy).pack(side='left')
        popup.wait_window()

    def _registrar_entrada_organista(self):
        self._popup_registrar_organista('organista_entrada')

    def _registrar_saida_organista(self):
        self._popup_registrar_organista('organista_saida')

    def _on_click_organista_toggle(self):
        """Decide automaticamente se o próximo evento deve ser entrada ou saída com base nos registos de hoje."""
        try:
            entradas = self.db.obter_eventos_por_tipo('organista_entrada') or []
            saidas = self.db.obter_eventos_por_tipo('organista_saida') or []
            # Se houve mais entradas do que saídas, aguardamos uma saída
            if len(entradas) > len(saidas):
                ev = 'organista_saida'
            else:
                ev = 'organista_entrada'
        except Exception:
            ev = 'organista_entrada'
        self._popup_registrar_organista(ev)

    def _update_organista_button_state(self):
        """Atualiza o rótulo do botão do organista consoante os eventos registados hoje."""
        try:
            entradas = self.db.obter_eventos_por_tipo('organista_entrada') or []
            saidas = self.db.obter_eventos_por_tipo('organista_saida') or []
            # Se houve mais entradas do que saídas, o próximo passo é registar a saída
            if len(entradas) > len(saidas):
                texto = "❌ Registar Saída Organista"
                bg = "#f56565"        # vermelho
                active_bg = "#c53030"
            else:
                texto = "✓ Registar Entrada Organista"
                bg = "#48bb78"        # verde
                active_bg = "#38a169"
            try:
                self.btn_organista.config(text=texto, bg=bg, activebackground=active_bg)
            except Exception:
                pass
        except Exception:
            pass

    def _mostrar_detalhes(self, event=None):
        # mostra um popup com os detalhes da linha (especialmente as anotações completas)
        sel = self.tree.selection()
        if not sel:
            return
        item = sel[0]
        vals = self.tree.item(item, "values")
        # assumimos que anotacoes é a última coluna
        anot = vals[-1] if vals and len(vals) > 0 else ""

        popup = tk.Toplevel(self.root)
        popup.title("Anotações")
        popup.geometry("500x300")
        popup.transient(self.root)
        txt = scrolledtext.ScrolledText(popup, font=("Segoe UI", 10), wrap=tk.WORD)
        txt.pack(expand=True, fill="both", padx=10, pady=10)
        txt.insert("1.0", anot)
        txt.config(state="disabled")
        ttk.Button(popup, text="Fechar", command=popup.destroy).pack(pady=(0, 10))

    def _proximo_numero_bilhete(self):
        ultimo = self.db.ultimo_numero_bilhete()
        ano = datetime.now().year
        if ultimo and isinstance(ultimo, str) and ultimo.startswith(f"IG{ano}-"):
            try:
                n = int(ultimo.split("-")[1])
            except Exception:
                n = 0
            return n + 1
        return 1

    def _on_nacionalidade_change(self, event=None):
        try:
            val = self.combo_nacionalidade.get()
        except Exception:
            return
        # se selecionado 'Outros', mostrar entrada manual
        if val and val.lower() == 'outros':
            # inserir a entrada manual na linha definida
            try:
                # mostrar rótulo e entrada na mesma linha
                self.lbl_manual_nacionalidade.grid(row=self._manual_row, column=0, sticky='w', pady=12, padx=(0, 10))
                self.entry_manual_nacionalidade.grid(row=self._manual_row, column=1, sticky='ew', pady=12)
            except Exception:
                pass
        else:
            # esconder a entrada manual
            try:
                self.entry_manual_nacionalidade.grid_forget()
                self.lbl_manual_nacionalidade.grid_forget()
                self.manual_nacionalidade_var.set('')
            except Exception:
                pass

    def _on_fatura_change(self, event=None):
        try:
            val = self.combo_fatura.get()
        except Exception:
            return
        if val and val.lower() == 'sim':
            try:
                self.lbl_contribuinte.grid(row=self._contrib_row, column=0, sticky='w', pady=12, padx=(0, 10))
                self.entry_contribuinte.grid(row=self._contrib_row, column=1, sticky='ew', pady=12)
            except Exception:
                pass
        else:
            try:
                self.entry_contribuinte.grid_forget()
                self.lbl_contribuinte.grid_forget()
                self.contribuinte_var.set('')
            except Exception:
                pass

    def guardar_registo(self):
        if self.dia_fechado:
            messagebox.showwarning("Aviso", "O dia já está fechado! Não é possível registar bilhetes.")
            return

        nacionalidade = self.combo_nacionalidade.get()
        # Se selecionado 'Outros', exigir valor manual e usá-lo
        try:
            if nacionalidade and nacionalidade.lower() == 'outros':
                manual_val = self.manual_nacionalidade_var.get().strip()
                if not manual_val:
                    messagebox.showwarning("Aviso", "Por favor especifique a nacionalidade quando selecionar 'Outros'.")
                    return
                nacionalidade = manual_val
        except Exception:
            pass
        metodo_pagamento = self.combo_pagamento.get()
        fatura = self.combo_fatura.get()
        # contribuinte is required when fatura == 'Sim'
        contribuinte = None
        try:
            if fatura and fatura.lower() == 'sim':
                contribuinte = self.contribuinte_var.get().strip()
                if not contribuinte:
                    messagebox.showwarning("Aviso", "Por favor insira o Nº Contribuinte quando selecionar 'Sim' em Recibo.")
                    return
            else:
                # ensure empty when not required
                contribuinte = None
        except Exception:
            contribuinte = None
        try:
            quantidade = int(self.spin_quantidade.get())
            if quantidade <= 0:
                raise ValueError()
        except Exception:
            messagebox.showwarning("Aviso", "Quantidade inválida.")
            return

        # ScrolledText: obter todo o texto (multi-linha)
        try:
            anotacoes = self.entry_anotacoes.get("1.0", "end").strip() or None
        except Exception:
            # se por alguma razão não for ScrolledText (compatibilidade), tentar Entry
            anotacoes = getattr(self, 'entry_anotacoes').get().strip() or None

        # preparar números dos bilhetes (ainda não gravar — pedir pagamento primeiro)
        proximo = self._proximo_numero_bilhete()
        ano = datetime.now().year
        data_hora = agora_str()
        bilhetes = [f"IG{ano}-{proximo + i}" for i in range(quantidade)]

        # limpar campos e atualizar
        self.combo_nacionalidade.set("Português")
        # esconder/limpar entrada manual se visível
        try:
            self.entry_manual_nacionalidade.grid_forget()
            self.manual_nacionalidade_var.set("")
        except Exception:
            pass
        self.combo_pagamento.set("Dinheiro")
        self.combo_fatura.set("Não")
        self.entry_contribuinte.delete(0, tk.END)
        # limpar anotacoes
        # limpar anotacoes (ScrolleText)
        try:
            self.entry_anotacoes.delete("1.0", tk.END)
        except Exception:
            try:
                self.entry_anotacoes.delete(0, tk.END)
            except Exception:
                pass
        self.spin_quantidade.set(1)

        # Após preparar os bilhetes, dependendo do método de pagamento:
        # - se for 'Dinheiro' abrir popup para introduzir valor recebido e calcular troco
        # - se for outro método (ex. cartão/multibanco) gravar diretamente e gerar o PDF
        try:
            total_price = len(bilhetes) * getattr(self, 'ticket_price', TICKET_PRICE)
            metodo_norm = (metodo_pagamento or "").strip().lower()
            # Se quantidade > 1, vamos criar apenas um bilhete que indica a quantidade
            agrupado = len(bilhetes) > 1
            if metodo_norm and metodo_norm != 'dinheiro':
                # pagamento por cartão: gravar registos e gerar PDF sem pedir valor recebido
                try:
                    if agrupado:
                        # gravar individualmente cada número no BD, mas imprimir apenas um bilhete com a quantidade
                        # garantir que gravamos o preço unitário do bilhete para cada registo
                        try:
                            self.db._pending_preco = getattr(self, 'ticket_price', TICKET_PRICE)
                        except Exception:
                            pass
                        try:
                            self.db._pending_preco = getattr(self, 'ticket_price', TICKET_PRICE)
                        except Exception:
                            pass
                        try:
                            self.db._pending_preco = getattr(self, 'ticket_price', TICKET_PRICE)
                        except Exception:
                            pass
                        for numero in bilhetes:
                            try:
                                self.db.inserir_registo(data_hora, self.assistente, nacionalidade, numero, metodo_pagamento, fatura, contribuinte, anotacoes)
                            except Exception:
                                pass
                        try:
                            delattr(self.db, '_pending_preco')
                        except Exception:
                            try:
                                del self.db._pending_preco
                            except Exception:
                                pass
                        try:
                            delattr(self.db, '_pending_preco')
                        except Exception:
                            try:
                                del self.db._pending_preco
                            except Exception:
                                pass
                        try:
                            delattr(self.db, '_pending_preco')
                        except Exception:
                            try:
                                del self.db._pending_preco
                            except Exception:
                                pass
                        # Anexar quantidade nas anotações do primeiro bilhete para exibição
                        try:
                            self.db.atualizar_anotacoes_por_numero(bilhetes[0], f"Qtd:{len(bilhetes)}")
                        except Exception:
                            pass
                        try:
                            self.atualizar_tabela()
                            self._atualizar_status()
                        except Exception:
                            pass
                        try:
                            # mensagem de sucesso (listar primeiros/mostrar contagem)
                            messagebox.showinfo("Sucesso", f"Foram registados {len(bilhetes)} bilhetes:\n{', '.join(bilhetes)}\n\nSerá impresso 1 bilhete com quantidade {len(bilhetes)}.")
                            self._set_status(f"{len(bilhetes)} bilhetes registados. Impresso 1 bilhete com quantidade.")
                        except Exception:
                            pass
                        try:
                            imprimir_bilhetes_multiplo_pdf([bilhetes[0]], data_hora, self.assistente, metodo_pagamento=metodo_pagamento, quantidade=len(bilhetes), preco=getattr(self, 'ticket_price', TICKET_PRICE))
                        except Exception as e:
                            print(f"Erro ao gerar/mandar imprimir PDF dos bilhetes: {e}")
                    else:
                        # única entrada: inserir normalmente
                        try:
                            self.db._pending_preco = getattr(self, 'ticket_price', TICKET_PRICE)
                        except Exception:
                            pass
                        for numero in bilhetes:
                            try:
                                self.db.inserir_registo(data_hora, self.assistente, nacionalidade, numero, metodo_pagamento, fatura, contribuinte, anotacoes)
                            except Exception:
                                pass
                        try:
                            delattr(self.db, '_pending_preco')
                        except Exception:
                            try:
                                del self.db._pending_preco
                            except Exception:
                                pass
                        try:
                            self.atualizar_tabela()
                            self._atualizar_status()
                        except Exception:
                            pass
                        try:
                            messagebox.showinfo("Sucesso", f"Foram registados {len(bilhetes)} bilhete(s):\n{', '.join(bilhetes)}")
                            self._set_status(f"{len(bilhetes)} bilhete(s) registado(s).")
                        except Exception:
                            pass
                        try:
                            imprimir_bilhetes_multiplo_pdf(bilhetes, data_hora, self.assistente, metodo_pagamento=metodo_pagamento, preco=getattr(self, 'ticket_price', TICKET_PRICE))
                        except Exception as e:
                            print(f"Erro ao gerar/mandar imprimir PDF dos bilhetes: {e}")
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao gravar registos:\n{e}")
            else:
                # pagamento em numerário: pedir valor recebido via popup
                # passar informação de quantidade para que o popup grave agrupado se necessário
                self._pedir_pagamento_e_imprimir(bilhetes, data_hora, total_price,
                                                 nacionalidade, metodo_pagamento, fatura, contribuinte, anotacoes, quantidade=len(bilhetes))
        except Exception as e:
            print(f"Erro ao iniciar fluxo de pagamento: {e}")

    def _pedir_pagamento_e_imprimir(self, bilhetes, data_hora, total_price,
                                    nacionalidade=None, metodo_pagamento=None, fatura=None, contribuinte=None, anotacoes=None, quantidade=1):
        """Abre um popup modal para introduzir o valor recebido pelo cliente, mostra o troco e confirma antes de gerar o PDF."""
        popup = tk.Toplevel(self.root)
        popup.title("Pagamento")
        popup.geometry("360x180")
        popup.transient(self.root)
        popup.grab_set()

        tk.Label(popup, text=f"Total a Pagar: €{total_price:.2f}", font=AF(11, "bold")).pack(pady=(12, 6))

        entry_frame = tk.Frame(popup)
        entry_frame.pack(pady=(6, 6))
        tk.Label(entry_frame, text="Valor Recebido: €", font=AF(10)).pack(side="left")
        recebido_var = tk.StringVar()
        recebido_entry = ttk.Entry(entry_frame, textvariable=recebido_var, width=12, font=AF(10))
        recebido_entry.pack(side="left")
        recebido_entry.focus()

        troco_label = tk.Label(popup, text="Troco: €0.00", font=AF(10))
        troco_label.pack(pady=(6, 8))

        def _update_troco(*args):
            s = recebido_var.get().strip()
            try:
                val = float(s.replace(',', '.')) if s else 0.0
                troco = val - float(total_price)
                troco_label.config(text=f"Troco: €{troco:.2f}")
            except Exception:
                troco_label.config(text="Troco: —")

        recebido_var.trace_add('write', _update_troco)

        btn_frame = tk.Frame(popup)
        btn_frame.pack(pady=(6, 8))

        def confirmar_pagamento():
            s = recebido_var.get().strip()
            try:
                received = float(s.replace(',', '.'))
            except Exception:
                messagebox.showwarning("Aviso", "Introduza um valor recebido válido (ex.: 10.00)")
                return
            if received < total_price:
                if not messagebox.askyesno("Valor Inferior", "O valor recebido é inferior ao total. Deseja continuar mesmo assim?"):
                    return
            try:
                troco = received - total_price
            except Exception:
                troco = 0.0
            # mostrar troco final antes de prosseguir
            messagebox.showinfo("Troco", f"Troco a entregar: €{troco:.2f}")
            popup.destroy()
            # após confirmação, gravar os registos no BD, atualizar UI e gerar PDF
            try:
                # gravar individualmente cada número no BD (mesmo que a impressão seja agrupada)
                try:
                    try:
                        self.db._pending_preco = getattr(self, 'ticket_price', TICKET_PRICE)
                    except Exception:
                        pass
                    for numero in bilhetes:
                        try:
                            self.db.inserir_registo(data_hora, self.assistente, nacionalidade, numero, metodo_pagamento, fatura, contribuinte, anotacoes)
                        except Exception:
                            # continuar a tentar inserir outros bilhetes
                            pass
                    try:
                        delattr(self.db, '_pending_preco')
                    except Exception:
                        try:
                            del self.db._pending_preco
                        except Exception:
                            pass
                    # Anexar quantidade nas anotações do primeiro bilhete para exibição
                    try:
                        if quantidade and int(quantidade) > 1:
                            self.db.atualizar_anotacoes_por_numero(bilhetes[0], f"Qtd:{int(quantidade)}")
                    except Exception:
                        pass
                except Exception:
                    pass
                # limpar campos e atualizar
                try:
                    self.combo_nacionalidade.set("Português")
                    self.entry_manual_nacionalidade.grid_forget()
                    self.manual_nacionalidade_var.set("")
                except Exception:
                    pass
                try:
                    self.combo_pagamento.set("Dinheiro")
                    self.combo_fatura.set("Não")
                    self.entry_contribuinte.delete(0, tk.END)
                except Exception:
                    pass
                try:
                    self.entry_anotacoes.delete("1.0", tk.END)
                except Exception:
                    try:
                        self.entry_anotacoes.delete(0, tk.END)
                    except Exception:
                        pass
                try:
                    self.spin_quantidade.set(1)
                except Exception:
                    pass

                # atualizar tabela e estatísticas
                try:
                    self.atualizar_tabela()
                    self._atualizar_status()
                except Exception:
                    pass

                # informar sucesso e gerar PDF
                try:
                    # mensagem de sucesso e resumo
                    if quantidade and int(quantidade) > 1:
                        messagebox.showinfo("Sucesso", f"Foram registados {int(quantidade)} bilhetes:\n{', '.join(bilhetes)}\n\nSerá impresso 1 bilhete com quantidade {int(quantidade)}.")
                        self._set_status(f"{int(quantidade)} bilhetes registados. Impresso 1 bilhete com quantidade.")
                    else:
                        messagebox.showinfo("Sucesso", f"Foram registados {len(bilhetes)} bilhete(s):\n{', '.join(bilhetes)}")
                        self._set_status(f"{len(bilhetes)} bilhete(s) registado(s).")
                except Exception:
                    pass
                try:
                    if quantidade and int(quantidade) > 1:
                        imprimir_bilhetes_multiplo_pdf([bilhetes[0]], data_hora, self.assistente, metodo_pagamento=metodo_pagamento, recebido=received, troco=troco, quantidade=quantidade, preco=getattr(self, 'ticket_price', TICKET_PRICE))
                    else:
                        imprimir_bilhetes_multiplo_pdf(bilhetes, data_hora, self.assistente, metodo_pagamento=metodo_pagamento, recebido=received, troco=troco, preco=getattr(self, 'ticket_price', TICKET_PRICE))
                except Exception as e:
                    print(f"Erro ao gerar/mandar imprimir PDF dos bilhetes: {e}")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao gravar registos após confirmação do pagamento:\n{e}")

        def cancelar():
            popup.destroy()

        ttk.Button(btn_frame, text="Confirmar", command=confirmar_pagamento).pack(side="left", padx=8)
        ttk.Button(btn_frame, text="Cancelar", command=cancelar).pack(side="left")

        popup.wait_window()

    def atualizar_tabela(self):
        # refresh table with today's data
        for ch in self.tree.get_children():
            self.tree.delete(ch)
        dados = self.db.obter_registos_do_dia()
        for idx, row in enumerate(dados):
            tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
            # ensure data_hora is string
            linha = list(row)
            linha[0] = str(linha[0])
            self.tree.insert("", "end", values=linha, tags=(tag,))
        self._atualizar_status()
        self._atualizar_estatisticas()

    def pesquisar_bilhete(self):
        termo = self.entry_search.get().strip()
        if not termo:
            self.atualizar_tabela()
            return
        for ch in self.tree.get_children():
            self.tree.delete(ch)
        dados = self.db.procurar_por_bilhete(termo)
        for idx, row in enumerate(dados):
            tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
            linha = list(row)
            linha[0] = str(linha[0])
            self.tree.insert("", "end", values=linha, tags=(tag,))
        self._set_status(f"Filtro: '{termo}' ({len(dados)} resultados)")

    def fechar_dia(self):
        if self.dia_fechado:
            messagebox.showinfo("Informação", "O dia já está fechado!")
            return
        if not messagebox.askyesno("Fechar o Dia", "Tem a certeza que deseja fechar o dia?"):
            return
        # pedir anotações finais (opcional)
        popup = tk.Toplevel(self.root)
        popup.title("Anotações Finais (opcional)")
        popup.geometry("500x300")
        popup.transient(self.root)
        popup.grab_set()
        tk.Label(popup, text="Anotações finais (opcional):", font=("Segoe UI", 10, "bold")).pack(anchor='w', padx=10, pady=(10, 0))
        txt_final = scrolledtext.ScrolledText(popup, font=("Segoe UI", 10), wrap=tk.WORD, height=10)
        txt_final.pack(expand=True, fill='both', padx=10, pady=10)

        def confirmar_fecho():
            try:
                self.final_notes = txt_final.get("1.0", "end").strip() or None
            except Exception:
                self.final_notes = None
            popup.destroy()
            # marcar dia fechado e seguir
            self.dia_fechado = True
            # desativar inputs
            for w in [self.combo_nacionalidade, self.combo_pagamento, self.combo_fatura,
                      self.entry_contribuinte, self.entry_anotacoes, self.entry_manual_nacionalidade, self.spin_quantidade, getattr(self, 'spin_reg_nao_entraram', None), getattr(self, 'btn_reg_nao', None)]:
                try:
                    w.config(state="disabled")
                except Exception:
                    pass
            self.gerar_excel()
            try:
                # atualizar o relatório horário mestre ao fechar o dia
                self.gerar_relatorio_horario()
            except Exception:
                pass
            self.criar_backup()
            self._set_status("Dia fechado. Relatórios gerados.")

        def cancelar_fecho():
            popup.destroy()
            return

        btns = tk.Frame(popup)
        btns.pack(pady=(0, 10))
        ttk.Button(btns, text="Confirmar Fecho", command=confirmar_fecho).pack(side='left', padx=8)
        ttk.Button(btns, text="Cancelar", command=cancelar_fecho).pack(side='left')
        popup.wait_window()
        # defensive default
        if not hasattr(self, 'final_notes'):
            self.final_notes = None

    # --------------------------
    # BACKUP E RELATÓRIOS
    # --------------------------
    def criar_backup(self):
        hoje = hoje_str()
        pasta_backup = "backups"
        os.makedirs(pasta_backup, exist_ok=True)
        backup_nome = os.path.join(pasta_backup, f"backup_bilhetes_{hoje}.db")
        try:
            if os.path.exists(self.db.path):
                shutil.copy(self.db.path, backup_nome)
                messagebox.showinfo("Backup Criado", f"Cópia de segurança criada em:\n{backup_nome}")
            else:
                messagebox.showwarning("Aviso", "Ficheiro de BD não encontrado para backup.")
        except Exception as e:
            messagebox.showerror("Erro no Backup", f"Falha ao criar cópia de segurança:\n{e}")

    def gerar_excel(self):
        hoje = hoje_str()
        pasta = os.path.join("relatorios", hoje)
        os.makedirs(pasta, exist_ok=True)
        dados = self.db.obter_registos_do_dia()
        if not dados:
            messagebox.showinfo("Sem Dados", "Não existem registos para hoje.")
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Bilhetes do Dia"
        cabecalho = ["Data/Hora", "Assistente", "Nacionalidade", "Número Bilhete", "Método Pagamento", "Recibo", "Contribuinte", "Preço", "Anotações"]
        ws.append(cabecalho)
        for col_num, _ in enumerate(cabecalho, 1):
            ws[f"{get_column_letter(col_num)}1"].font = Font(bold=True)
        for row in dados:
            # row now includes preco before anotacoes
            ws.append([str(x) if x is not None else "" for x in row])
        # aplicar wrap na coluna 'Anotações' (última coluna)
        try:
            anot_col = len(cabecalho)
            for cell in ws[get_column_letter(anot_col)]:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        except Exception:
            pass
        total = len(dados)
        ws.append([])
        ws.append(["Total de Bilhetes Vendidos:", total])
        # Calcular e adicionar totais monetários ao Excel
        try:
            cash_amount = 0.0
            card_amount = 0.0
            for row in dados:
                metodo = (row[4] or "").strip().lower()
                try:
                    preco_val = float(row[7]) if row[7] is not None else float(getattr(self, 'ticket_price', TICKET_PRICE))
                except Exception:
                    preco_val = float(getattr(self, 'ticket_price', TICKET_PRICE))
                if metodo == 'dinheiro':
                    cash_amount += preco_val
                elif metodo.startswith('cart') or 'multibanco' in metodo or 'cartão' in metodo:
                    card_amount += preco_val
            numerario_total = INITIAL_CASH + cash_amount
            caixa_total = numerario_total + card_amount
            ws.append(["Numerário:", f"€{numerario_total:.2f}"])
            ws.append(["Multibanco:", f"€{card_amount:.2f}"])
            ws.append(["Caixa total:", f"€{caixa_total:.2f}"])
        except Exception:
            # não impedir criação do Excel se falhar o cálculo
            pass
        # Incluir registos 'Não Entraram' (horas) na folha, se existirem eventos para hoje
        try:
            eventos = self.db.obter_eventos_por_tipo('nao_entraram')
            if eventos:
                ws.append([])
                ws.append(["Registos 'Não Entraram' (horas):"])
                # adicionar cabeçalho simples: Hora, Assistente, Quantidade
                ws.append(["Hora", "Assistente", "Quantidade"])
                # eventos is list of (id, timestamp, event_type, count, assistente, notes)
                for ev in reversed(eventos):
                    _, ts, _, cnt, assist, notes = ev
                    try:
                        hora = ts.split(' ')[1]
                    except Exception:
                        hora = ts
                    ws.append([hora, assist or "", cnt or ""])
                # adicionar total de pessoas que não entraram
                try:
                    total_nao_entraram = sum(int(ev[3] or 0) for ev in eventos)
                    ws.append([])
                    ws.append(["Total Não Entraram:", total_nao_entraram])
                except Exception:
                    pass
        except Exception:
            pass
        # Incluir registos de entrada/saída do organista
        try:
            entradas = self.db.obter_eventos_por_tipo('organista_entrada') or []
            saidas = self.db.obter_eventos_por_tipo('organista_saida') or []
            organista_events = list(entradas) + list(saidas)
            # ordenar por timestamp asc (opcional) - eventos já filtrados por dia na query
            if organista_events:
                ws.append([])
                ws.append(["Registos Organista:"])
                ws.append(["Hora", "Registrado Por", "Evento", "Organista", "Notas"])
                # eventos is list of (id, timestamp, event_type, count, assistente, notes)
                for ev in reversed(organista_events):
                    _, ts, ev_type, cnt, registrador, notes = ev
                    try:
                        hora = ts.split(' ')[1]
                    except Exception:
                        hora = ts
                    evento_nome = 'Entrada' if ev_type == 'organista_entrada' else 'Saída' if ev_type == 'organista_saida' else ev_type
                    organista_nome = ''
                    notas_extra = ''
                    try:
                        if notes:
                            parts = str(notes).split('|', 1)
                            organista_nome = parts[0]
                            if len(parts) > 1:
                                notas_extra = parts[1]
                    except Exception:
                        organista_nome = str(notes)
                    ws.append([hora, registrador or "", evento_nome, organista_nome or "", notas_extra or ""])
        except Exception:
            pass
        # incluir anotações finais no Excel: uma linha após os totais e, opcionalmente, numa aba separada
        try:
            if hasattr(self, 'final_notes') and self.final_notes:
                # adicionar linha simples abaixo
                ws.append([])
                ws.append(["Anotações Finais:", self.final_notes])
                # criar página separada com notas (maior legibilidade)
                notas_ws = wb.create_sheet(title="Notas Finais")
                notas_ws.append(["Anotações Finais"])
                # quebrar em linhas para inserir na sheet
                for ln in (self.final_notes or '').split('\n'):
                    notas_ws.append([ln])
                # aplicar wrap na primeira coluna
                for row in notas_ws.iter_rows(min_row=1, max_col=1):
                    for cell in row:
                        try:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                        except Exception:
                            pass
        except Exception:
            pass
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            # limitar largura máxima razoável
            width = min(max_len + 5, 100)
            ws.column_dimensions[get_column_letter(col[0].column)].width = width
        filename = os.path.join(pasta, f"Bilhetes_{hoje}.xlsx")
        try:
            wb.save(filename)
            messagebox.showinfo("Excel Gerado", f"Arquivo Excel criado: {filename}")
        except Exception as e:
            messagebox.showerror("Erro Excel", f"Falha ao criar Excel:\n{e}")

    def gerar_relatorio_horario(self):
        """Gera/atualiza `relatorios/estatisticas_horario.xlsx` com as informações pedidas pelo utilizador.

        Para cada hora onde existam registos, escreve uma linha com:
        - Dia (YYYY-MM-DD)
        - Intervalo (HH:00-HH:59)
        - Dia da semana (pt)
        - Assistente (nome mais frequente na hora)
        - Organista (S/N) — 'S' se existir evento 'organista_entrada' nessa hora
        - Nacionalidades base e respetivas quantidades (formatadas)
        - Outras nacionalidades e respetiva quantidade (formatadas)
        - Não pagantes por hora (registos com preco == 0)
        - Total de visitantes por hora (contagem de registos)

        No final adiciona linhas de total do dia: total não pagantes e total visitantes.
        """
        from collections import Counter
        from datetime import datetime
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        hoje = hoje_str()
        # mapa dia da semana em português
        weekdays_pt = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira", "Sábado", "Domingo"]

        pasta = os.path.join("relatorios")
        os.makedirs(pasta, exist_ok=True)
        caminho = os.path.join(pasta, "Estatísticas.xlsx")

        # definir nacionalidades base (consistente com a UI)
        BASE_NACIONALIDADES = ["Português", "Brasileiro", "Espanhol", "Inglês", "Francês", "Italiano", "Asiático", "Alemão"]
        base_lower = [b.lower() for b in BASE_NACIONALIDADES]

        # obter todos os registos do dia para agrupar por hora
        try:
            self.db.cursor.execute(
                "SELECT data_hora, assistente, nacionalidade, metodo_pagamento, preco FROM registos WHERE date(data_hora) = ? ORDER BY data_hora",
                (hoje,)
            )
            rows = self.db.cursor.fetchall()
        except Exception:
            rows = []

        # agrupar por hora
        hora_groups = {}
        for r in rows:
            try:
                dt = r[0]
                hh = dt[11:13]  # 'YYYY-MM-DD HH:MM:SS'
            except Exception:
                continue
            hora_groups.setdefault(hh, []).append({'data_hora': r[0], 'assistente': r[1], 'nacionalidade': r[2], 'metodo_pagamento': r[3], 'preco': r[4]})

        # organista: marcar horas onde existe evento de entrada
        organista_hours = set()
        try:
            eventos = self.db.obter_eventos_por_tipo('organista_entrada', hoje)
            for ev in eventos:
                ts = ev[1]  # timestamp
                try:
                    h = ts[11:13]
                    organista_hours.add(h)
                except Exception:
                    pass
        except Exception:
            pass

        # obter eventos 'nao_entraram' por hora (estes são os "não pagantes" solicitados agora)
        nao_entraram_by_hour = {}
        try:
            eventos_nao = self.db.obter_eventos_por_tipo('nao_entraram', hoje)
            for ev in eventos_nao:
                ts = ev[1]
                cnt = int(ev[3] or 0)
                try:
                    h = ts[11:13]
                    nao_entraram_by_hour[h] = nao_entraram_by_hour.get(h, 0) + cnt
                except Exception:
                    pass
        except Exception:
            nao_entraram_by_hour = {}

        # abrir ou criar workbook
        try:
            if os.path.exists(caminho):
                wb = load_workbook(caminho)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = 'Estatísticas'
                # título e metadados
                ws.append(["Estatísticas Horário"])

                # cabeçalho da tabela
                ws.append(["Dia", "Intervalo", "Dia da Semana", "Assistente", "Organista (S/N)", "Nacionalidades Base (contagens)", "Total Base", "Outras Nacionalidades (list)", "Total Outras", "Nao Pagantes Hora", "Total Visitantes Hora"])
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Estatísticas'
            ws.append(["Estatísticas Horário"])
            ws.append([""])
            ws.append(["Descrição das Colunas:"])
            ws.append(["(ver descrições no topo do ficheiro)"])
            ws.append([""])
            ws.append(["Dia", "Intervalo", "Dia da Semana", "Assistente", "Organista (S/N)", "Nacionalidades Base (contagens)", "Total Base", "Outras Nacionalidades (list)", "Total Outras", "Nao Pagantes Hora", "Total Visitantes Hora"])

        # remover linhas existentes do dia (para evitar duplicados ao re-fechar)
        try:
            to_delete = []
            for row_idx in range(1, ws.max_row + 1):
                try:
                    cell = ws.cell(row=row_idx, column=1).value
                    if cell == hoje:
                        to_delete.append(row_idx)
                except Exception:
                    pass
            for r in reversed(to_delete):
                ws.delete_rows(r, 1)
        except Exception:
            pass

        total_nao_pagantes_dia = sum(nao_entraram_by_hour.values())
        total_visitantes_dia = 0

        # identificar a linha do cabeçalho (primeira ocorrência do texto 'Intervalo')
        header_row = None
        for row_idx in range(1, ws.max_row + 1):
            try:
                if str(ws.cell(row=row_idx, column=1).value).strip().lower() == 'dia' and str(ws.cell(row=row_idx, column=2).value).strip().lower() in ('intervalo', 'intervalo'):
                    header_row = row_idx
                    break
            except Exception:
                continue
        if header_row is None:
            header_row = ws.max_row + 1
            ws.append(["Dia", "Intervalo", "Dia da Semana", "Assistente", "Organista (S/N)  ", "Nacionalidades Base (contagens)", "Total Base"  , "Outras Nacionalidades (list)", "Total Outras ", "Nao Pagantes Hora"  , "Total Visitantes Hora    "])

        # iterar apenas horas com registos
        for hh in sorted(hora_groups.keys()):
            group = hora_groups[hh]
            intervalo = f"{hh}:00-{hh}:59"
            dia_sem = ''
            try:
                dia_sem = weekdays_pt[datetime.strptime(hoje, "%Y-%m-%d").weekday()]
            except Exception:
                dia_sem = ''

            # assistente mais frequente na hora
            assistentes = [g['assistente'] or '' for g in group]
            assistente_counter = Counter(assistentes)
            assistente_top = ''
            if assistente_counter:
                assistente_top = assistente_counter.most_common(1)[0][0]

            # organista presente?
            organista_flag = 'S' if hh in organista_hours else 'N'

            # nacionalidades
            base_counts = Counter()
            outras_counts = Counter()
            for g in group:
                nat = (g.get('nacionalidade') or '').strip()
                if not nat:
                    continue
                if nat.lower() in base_lower:
                    idx = base_lower.index(nat.lower())
                    base_key = BASE_NACIONALIDADES[idx]
                    base_counts[base_key] += 1
                else:
                    outras_counts[nat] += 1

            total_base = sum(base_counts.values())
            total_outras = sum(outras_counts.values())

            # não pagantes por hora: usar eventos 'nao_entraram' agrupados por hora
            nao_pagantes_hora = nao_entraram_by_hour.get(hh, 0)
            total_visitantes_hora = len(group)

            total_visitantes_dia += total_visitantes_hora

            # formatar colunas de nacionalidades
            base_fmt = "; ".join([f"{k}: {v}" for k, v in base_counts.items()]) if base_counts else ""
            outras_fmt = "; ".join([f"{k}: {v}" for k, v in outras_counts.items()]) if outras_counts else ""

            ws.append([hoje, intervalo, dia_sem, assistente_top, organista_flag, base_fmt, total_base, outras_fmt, total_outras, nao_pagantes_hora, total_visitantes_hora])

        # totais do dia
        try:
            ws.append(['', '', '', '', '', '', '', '', '', total_nao_pagantes_dia, total_visitantes_dia])
            ws.append([])
        except Exception:
            pass

        # formatação: cabeçalho em negrito, larguras, alinhamentos e filtro
        try:
            # localizar a linha do cabeçalho novamente
            hrow = None
            for row_idx in range(1, ws.max_row + 1):
                try:
                    if str(ws.cell(row=row_idx, column=2).value).strip().lower() == 'intervalo':
                        hrow = row_idx
                        break
                except Exception:
                    continue
            if hrow is None:
                hrow = header_row

            header_font = Font(bold=True)
            for col in range(1, 12):
                try:
                    ws.cell(row=hrow, column=col).font = header_font
                    ws.cell(row=hrow, column=col).alignment = Alignment(horizontal='center', vertical='center')
                except Exception:
                    pass

            # congelar painéis após o cabeçalho
            try:
                ws.freeze_panes = ws.cell(row=hrow+1, column=1)
            except Exception:
                pass

            # ajustar larguras de colunas automaticamente com base no conteúdo
            from openpyxl.utils import get_column_letter
            try:
                last_row = ws.max_row
                col_count = 11
                max_widths = [0] * col_count
                for col in range(1, col_count + 1):
                    for row in range(hrow, last_row + 1):
                        try:
                            val = ws.cell(row=row, column=col).value
                            if val is None:
                                continue
                            text = str(val)
                            # considerar a linha mais longa em caso de wrap
                            lines = text.splitlines()
                            max_len = max((len(line) for line in lines), default=len(text))
                            if max_len > max_widths[col - 1]:
                                max_widths[col - 1] = max_len
                        except Exception:
                            continue

                # aplicar largura com padding e limites mínimos/máximos
                for i, mw in enumerate(max_widths, start=1):
                    try:
                        width = int(mw) + 2
                        if width < 8:
                            width = 8
                        if width > 100:
                            width = 100
                        ws.column_dimensions[get_column_letter(i)].width = width
                    except Exception:
                        pass
            except Exception:
                pass

            # wrap text for nationality columns
            for row in ws.iter_rows(min_row=hrow+1, min_col=6, max_col=9):
                for cell in row:
                    try:
                        cell.alignment = Alignment(wrap_text=True)
                    except Exception:
                        pass

            # aplicar autofiltro
            try:
                last_row = ws.max_row
                ws.auto_filter.ref = f"A{hrow}:K{last_row}"
            except Exception:
                pass
        except Exception:
            pass

        try:
            # --- Adicionar folha mensal com estatísticas do mês atual ---
            try:
                from collections import Counter
                import calendar
                from datetime import datetime

                mes_key = hoje[:7]  # 'YYYY-MM'
                # traduzir mês para nome em português (para exibição)
                try:
                    month_num = int(mes_key.split('-')[1])
                    months_pt = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
                    mes_nome_pt = months_pt[month_num - 1]
                except Exception:
                    mes_nome_pt = mes_key

                # buscar registos do mês
                try:
                    self.db.cursor.execute(
                        "SELECT data_hora, assistente, nacionalidade FROM registos WHERE substr(data_hora,1,7) = ? ORDER BY data_hora",
                        (mes_key,)
                    )
                    month_rows = self.db.cursor.fetchall()
                except Exception:
                    month_rows = []

                total_visitors_month = len(month_rows)

                # nao_entraram mensal (somar counts dos eventos)
                try:
                    self.db.cursor.execute(
                        "SELECT SUM(count) FROM eventos WHERE event_type = 'nao_entraram' AND substr(timestamp,1,7) = ?",
                        (mes_key,)
                    )
                    nao_sum = self.db.cursor.fetchone()[0]
                    total_nao_entraram_month = int(nao_sum or 0)
                except Exception:
                    total_nao_entraram_month = 0

                # nacionalidades do mês
                nat_counter = Counter()
                hours_set = set()
                days_set = set()
                visitors_by_hour_interval = {f"{h:02d}:00-{h:02d}:59": 0 for h in range(24)}
                visitors_by_weekday = Counter()
                hours_by_weekday_sets = {i: set() for i in range(7)}

                for r in month_rows:
                    dt = r[0]
                    try:
                        date_part = dt[:10]
                        hour_key = dt[:13]  # YYYY-MM-DD HH
                        hh = dt[11:13]
                    except Exception:
                        continue
                    nat = (r[2] or '').strip()
                    if nat:
                        nat_counter[nat] += 1
                    hours_set.add(hour_key)
                    days_set.add(date_part)
                    visitors_by_hour_interval[f"{int(hh):02d}:00-{int(hh):02d}:59"] += 1
                    # weekday: Monday=0
                    try:
                        wd = datetime.strptime(date_part, "%Y-%m-%d").weekday()
                        visitors_by_weekday[wd] += 1
                        hours_by_weekday_sets[wd].add(hour_key)
                    except Exception:
                        pass

                total_hours_with_visitors = len(hours_set)

                # organista hours across month
                organist_hours_month = set()
                try:
                    self.db.cursor.execute(
                        "SELECT timestamp FROM eventos WHERE event_type = 'organista_entrada' AND substr(timestamp,1,7) = ?",
                        (mes_key,)
                    )
                    evs = self.db.cursor.fetchall()
                    for e in evs:
                        try:
                            ts = e[0]
                            organist_hours_month.add(ts[:13])
                        except Exception:
                            pass
                except Exception:
                    pass

                # visitors with/without organist
                visitors_with_organist = 0
                visitors_without_organist = 0
                hours_with_organist = set()
                for r in month_rows:
                    try:
                        hour_key = r[0][:13]
                    except Exception:
                        continue
                    if hour_key in organist_hours_month:
                        visitors_with_organist += 1
                        hours_with_organist.add(hour_key)
                    else:
                        visitors_without_organist += 1

                hours_with_organist_count = len(hours_with_organist)
                hours_without_organist_count = total_hours_with_visitors - hours_with_organist_count

                visitors_per_hour_with_organist = visitors_with_organist / hours_with_organist_count if hours_with_organist_count else 0
                visitors_per_hour_without_organist = visitors_without_organist / hours_without_organist_count if hours_without_organist_count else 0

                avg_visitors_per_day = total_visitors_month / len(days_set) if len(days_set) else 0

                # visitors per weekday and hours per weekday and avg per hour per weekday
                weekdays_pt = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira", "Sábado", "Domingo"]
                visitors_by_weekday_named = {weekdays_pt[k]: v for k, v in visitors_by_weekday.items()}
                hours_by_weekday = {weekdays_pt[k]: len(s) for k, s in hours_by_weekday_sets.items()}
                avg_by_weekday = {}
                for k in range(7):
                    name = weekdays_pt[k]
                    hrs = hours_by_weekday.get(name, 0)
                    vis = visitors_by_weekday.get(k, 0)
                    avg_by_weekday[name] = (vis / hrs) if hrs else 0

                # criar ou substituir folha mensal
                mensal_name = f"Mensal_{mes_key}"
                if mensal_name in wb.sheetnames:
                    ws_m = wb[mensal_name]
                    wb.remove(ws_m)
                ws_m = wb.create_sheet(title=mensal_name)

                # preencher com métricas principais
                ws_m.append([f"Estatísticas Mensais - {mes_nome_pt}"])
                ws_m.append([""])
                ws_m.append(["Métrica", "Valor"])
                rows_metrics = [
                    ("Mês", mes_nome_pt),
                    ("Número de visitantes", total_visitors_month),
                    ("Número de não pagantes", total_nao_entraram_month),
                    ("Número de horas com visitas", total_hours_with_visitors),
                    ("Número de visitantes c/ organista", visitors_with_organist),
                    ("Número de horas com organista", hours_with_organist_count),
                    ("Número de visitantes s/ organista", visitors_without_organist),
                    ("Número de horas s/ organista", hours_without_organist_count),
                    ("Visitantes por hora (com organista)", round(visitors_per_hour_with_organist,2)),
                    ("Visitantes por hora (s/ organista)", round(visitors_per_hour_without_organist,2)),
                    ("Média de visitantes por dia", round(avg_visitors_per_day,2)),
                ]
                for m in rows_metrics:
                    ws_m.append(list(m))

                ws_m.append([""])

                # nacionalidades tabela
                ws_m.append(["Nacionalidade", "Quantidade"])
                for nat, cnt in nat_counter.most_common():
                    ws_m.append([nat, cnt])

                ws_m.append([""])

                # visitantes por hora (intervalos)
                ws_m.append(["Intervalo Hora", "Visitantes no Mês"])
                # listar apenas intervalos com visitantes (evitar linhas vazias)
                for k in sorted(visitors_by_hour_interval.keys()):
                    cnt = visitors_by_hour_interval[k]
                    if cnt > 0:
                        ws_m.append([k, cnt])

                ws_m.append([""])

                # visitantes por dia da semana
                ws_m.append(["Dia da Semana", "Visitantes", "Horas com visitas", "Média visitantes/hora"])
                for name in weekdays_pt:
                    vis = visitors_by_weekday.get(weekdays_pt.index(name), 0)
                    hrs = hours_by_weekday.get(name, 0)
                    avg = round(avg_by_weekday.get(name, 0), 2)
                    ws_m.append([name, vis, hrs, avg])

                # formatação simples para a folha mensal (negrito nos headers)
                try:
                    bold = Font(bold=True)
                    # aplicar negrito às primeiras linhas de cabeçalho
                    ws_m['A4'].font = bold
                    # bold nas cabeças das tabelas (aproximamos: procurar linhas com texto específico)
                    for row_idx in range(1, ws_m.max_row + 1):
                        try:
                            val = str(ws_m.cell(row=row_idx, column=1).value or '').lower()
                            if any(h in val for h in ('nacionalidade', 'intervalo hora', 'diasemana', 'métrica', 'intervalo')):
                                for col in range(1, ws_m.max_column + 1):
                                    try:
                                        ws_m.cell(row=row_idx, column=col).font = bold
                                    except Exception:
                                        pass
                        except Exception:
                            pass
                except Exception:
                    pass

                # formatação adicional: auto-ajustar colunas, freeze e autofiltro
                try:
                    from openpyxl.utils import get_column_letter
                    max_cols = ws_m.max_column
                    max_rows = ws_m.max_row
                    max_widths = [0] * max_cols
                    for col in range(1, max_cols + 1):
                        for row in range(1, max_rows + 1):
                            try:
                                v = ws_m.cell(row=row, column=col).value
                                if v is None:
                                    continue
                                s = str(v)
                                lines = s.splitlines()
                                l = max((len(x) for x in lines), default=len(s))
                                if l > max_widths[col - 1]:
                                    max_widths[col - 1] = l
                            except Exception:
                                continue
                    for i, mw in enumerate(max_widths, start=1):
                        try:
                            w = int(mw) + 2
                            if w < 8:
                                w = 8
                            if w > 120:
                                w = 120
                            ws_m.column_dimensions[get_column_letter(i)].width = w
                        except Exception:
                            pass
                    # wrap text for nationality column if present
                    for row in ws_m.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=max_cols):
                        for cell in row:
                            try:
                                cell.alignment = Alignment(wrap_text=True)
                            except Exception:
                                pass
                    # freeze top rows to keep title/legend visible
                    try:
                        ws_m.freeze_panes = ws_m['A6']
                    except Exception:
                        pass
                    # apply autofilter to the main tables where possible: try to find the 'Intervalo Hora' header row
                    try:
                        for r in range(1, max_rows + 1):
                            try:
                                if str(ws_m.cell(row=r, column=1).value).strip().lower() == 'intervalo hora':
                                    last = max_rows
                                    ws_m.auto_filter.ref = f"A{r}:B{last}"
                                    break
                            except Exception:
                                pass
                    except Exception:
                        pass
                except Exception:
                    pass
            except Exception:
                pass

            wb.save(caminho)
        except Exception:
            pass

    def gerar_pdf(self):
        hoje = hoje_str()
        if not REPORTLAB_AVAILABLE:
            messagebox.showwarning("Dependência em falta", "A biblioteca 'reportlab' não está instalada. Instale com: pip install reportlab")
            return
        pasta = os.path.join("relatorios", hoje)
        os.makedirs(pasta, exist_ok=True)
        dados = self.db.obter_registos_do_dia()
        if not dados:
            messagebox.showinfo("Sem Dados", "Não existem registos para hoje.")
            return

        filename = os.path.join(pasta, f"Bilhetes_{hoje}.pdf")
        pdf = SimpleDocTemplate(filename, pagesize=A4)
        elementos = []
        styles = getSampleStyleSheet()
        elementos.append(Paragraph(f"<b>Relatório de Bilhetes - {hoje}</b>", styles["Title"]))
        cabecalho = ["Data/Hora", "Assistente", "Nacionalidade", "Nº Bilhete", "Pagamento", "Recibo", "Contribuinte", "Preço", "Anotações"]
        tabela_dados = [cabecalho]
        for row in dados:
            r = list(row)
            try:
                anot_text = str(r[-1]) if r[-1] is not None else ""
            except Exception:
                anot_text = ""
            r[-1] = Paragraph(anot_text.replace('\n', '<br/>'), styles['BodyText'])
            tabela_dados.append(r)

        # Totais e summaries
        total = len(dados)
        empty_row = [""] * len(cabecalho)
        row_total = [""] * len(cabecalho)
        row_total[0] = "Total de Bilhetes Vendidos:"
        row_total[1] = total
        tabela_dados.append(empty_row)
        tabela_dados.append(row_total)

        # (campo 'Não Entraram' removido — não incluir essa linha no relatório)

        # Anotações finais (se existirem)
        try:
            if hasattr(self, 'final_notes') and self.final_notes:
                row_notes = [""] * len(cabecalho)
                row_notes[0] = Paragraph(self.final_notes.replace('\n', '<br/>'), styles['BodyText'])
                tabela_dados.append([""])  # spacer
                tabela_dados.append(row_notes)
        except Exception:
            pass

        t = Table(tabela_dados, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke)
        ]))
        elementos.append(t)
        try:
            pdf.build(elementos)
            messagebox.showinfo("PDF Gerado", f"Arquivo PDF criado: {filename}")
        except Exception as e:
            messagebox.showerror("Erro PDF", f"Falha ao criar PDF:\n{e}")

    # --------------------------
    # ESTATÍSTICAS E STATUS
    # --------------------------
    def _atualizar_estatisticas(self):
        # atualiza total e tabela por nacionalidade
        dados = self.db.obter_registos_do_dia()
        total = len(dados)
        self.lbl_total_today.config(text=f"Total de bilhetes hoje: {total}")

        # sumarizar por nacionalidade
        summary = {}
        for row in dados:
            nat = row[2] or "Outros"
            summary[nat] = summary.get(nat, 0) + 1

        # calcular totais por método de pagamento usando o preço armazenado por registo
        try:
            cash_amount = 0.0
            card_amount = 0.0
            for row in dados:
                metodo = (row[4] or "").strip().lower()
                try:
                    preco_val = float(row[7]) if row[7] is not None else float(getattr(self, 'ticket_price', TICKET_PRICE))
                except Exception:
                    preco_val = float(getattr(self, 'ticket_price', TICKET_PRICE))
                if metodo == 'dinheiro':
                    cash_amount += preco_val
                elif metodo.startswith('cart') or 'multibanco' in metodo or 'cartão' in metodo:
                    card_amount += preco_val
        except Exception:
            cash_amount = 0.0
            card_amount = 0.0

        # Numerário deve incluir o valor inicial da caixa
        numerario_total = INITIAL_CASH + cash_amount
        # Caixa total inclui numerário (com caixa inicial) e também o multibanco
        caixa_total = numerario_total + card_amount

        # atualizar rótulos de valores monetários
        try:
            self.lbl_numerario.config(text=f"Numerário: €{numerario_total:.2f}")
            self.lbl_multibanco.config(text=f"Multibanco: €{card_amount:.2f}")
            self.lbl_caixa_total.config(text=f"Caixa total: €{caixa_total:.2f}")
        except Exception:
            pass

        # limpar lista
        for ch in self.lst_nacionalidades.get_children():
            self.lst_nacionalidades.delete(ch)
        for nat, cnt in sorted(summary.items(), key=lambda x: x[1], reverse=True):
            self.lst_nacionalidades.insert("", "end", values=(nat, cnt))

    def _atualizar_status(self):
        # atualiza tabela e estatísticas
        self._set_status("Sistema pronto")
        self._atualizar_estatisticas()

    def _set_status(self, texto, timeout_ms=5000):
        self.status_var.set(texto)
        if timeout_ms:
            self.root.after(timeout_ms, lambda: self.status_var.set("Sistema pronto - Aguardando ações"))

    # --------------------------
    # ENCERRAMENTO
    # --------------------------
    def _on_close(self):
        if messagebox.askokcancel("Sair", "Deseja sair da aplicação?"):
            try:
                self.db.fechar()
            except Exception:
                pass
            self.root.destroy()
            # exit cleanly (helps if run from a double-click)
            try:
                sys.exit(0)
            except Exception:
                pass


# ==========================
# EXECUÇÃO
# ==========================
if __name__ == "__main__":
    JanelaLogin()